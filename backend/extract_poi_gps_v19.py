"""
Extract POI GPS coordinates from PortugalVivo_BaseDados_POI_v19.xlsx.

Reads every POI sheet, detects columns dynamically, validates coordinates
against the Portugal bounding box (continent + Açores + Madeira), and emits
a normalised JSON file consumable by `apply_poi_gps_v19.py`.

Output: backend/data/poi_gps_v19.json

Usage:
    python extract_poi_gps_v19.py
    python extract_poi_gps_v19.py --excel backend/data/PortugalVivo_BaseDados_POI_v19.xlsx
"""
from __future__ import annotations

import argparse
import json
import logging
import re
import sys
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

import openpyxl

logging.basicConfig(level=logging.INFO, format="%(message)s")
log = logging.getLogger(__name__)

DEFAULT_EXCEL = Path(__file__).parent / "data" / "PortugalVivo_BaseDados_POI_v19.xlsx"
DEFAULT_OUT = Path(__file__).parent / "data" / "poi_gps_v19.json"
DEFAULT_MISSING_OUT = Path(__file__).parent / "data" / "poi_gps_v19_missing.json"

# Sheets that aren't POI lists (system, navigation, indexes).
SYSTEM_SHEETS = {
    "🔧 MongoDB Schema",
    "🔧 Ingestão Config",
    "📊 Dashboard",
    "Categorias",
    "🗺 Guia do Viajante",
    "🚌 Transportes",
    "Grande Expedição 2026",
    "Música Tradicional",
    "Entidades e Operadores",
    "Rotas Temáticas",
    "Base de Dados POI",
}

# Sheet → app category (mirrors import_excel_v19_smart.SHEET_TO_CATEGORY).
SHEET_TO_CATEGORY = {
    "Percursos Pedestres": "percursos",
    "Praias Fluviais": "piscinas",
    "Termas e Banhos": "termas",
    "Cascatas e Poços Naturais": "cascatas",
    "Miradouros Portugal": "miradouros",
    "Castelos": "arqueologia",
    "Palácios e Solares": "arqueologia",
    "Museus": "arte",
    "Tabernas Históricas": "gastronomia",
    "Restaurantes e Gastronomia": "gastronomia",
    "Mercados e Feiras": "gastronomia",
    "Produtores DOP e Locais": "produtos",
    "Festas e Romarias": "festas",
    "Festivais de Música": "festas",
    "Aventura e Natureza": "aventura",
    "Natureza Especializada": "areas_protegidas",
    "Surf": "surf",
    "Ecovias e Passadiços": "percursos",
    "Praias Bandeira Azul": "piscinas",
    "Património Ferroviário": "arqueologia",
    "Arte Urbana e Intervenção": "arte",
    "Moinhos e Azenhas": "saberes",
    "Arqueologia, Geologia e Mineral": "arqueologia",
    "Parques de Campismo": "aventura",
    "Pousadas de Juventude": "aventura",
    "Flora Autóctone": "areas_protegidas",
    "Fauna Autóctone": "areas_protegidas",
    "Flora Botânica": "areas_protegidas",
    "Biodiversidade | Avistamentos": "areas_protegidas",
    "Barragens e Albufeiras": "piscinas",
    "Ofícios e Artesanato": "saberes",
    "Faróis": "miradouros",
    "Alojamentos Rurais": "aldeias",
    "Agroturismo e Enoturismo": "gastronomia",
    "Pérolas de Portugal": "aldeias",
    "Sopas Típicas": "gastronomia",
    "Pratos Típicos": "gastronomia",
    "Doçaria Regional": "gastronomia",
    "Agentes Turísticos": "rotas",
}

REGION_MAP = {
    "norte": "norte",
    "centro": "centro",
    "lisboa": "lisboa",
    "lisboa e vale do tejo": "lisboa",
    "area metropolitana de lisboa": "lisboa",
    "alentejo": "alentejo",
    "alto alentejo": "alentejo",
    "baixo alentejo": "alentejo",
    "algarve": "algarve",
    "acores": "acores",
    "açores": "acores",
    "madeira": "madeira",
    "minho": "norte",
    "douro": "norte",
    "tras-os-montes": "norte",
    "beira": "centro",
    "beira litoral": "centro",
    "beira interior": "centro",
    "estremadura": "lisboa",
    "ribatejo": "lisboa",
    "estremadura e ribatejo": "lisboa",
    "portugal continental": "portugal",
    "arquipelago dos acores": "acores",
    "arquipelago da madeira": "madeira",
}


def infer_region_from_coords(lat: float, lng: float) -> str:
    """Heuristic NUTS II classification by coordinates. Used when region text is generic."""
    # Açores (8 islands, lng -31 to -24, lat 36.5-39.8)
    if -32 <= lng <= -24 and 36 <= lat <= 40:
        return "acores"
    # Madeira (lng -17.3 to -16.2, lat 32.4-33.2)
    if -18 <= lng <= -16 and 32 <= lat <= 33.5:
        return "madeira"
    # Continental Portugal: split by latitude bands.
    if lat >= 41.0:
        return "norte"
    if lat >= 39.7:
        return "centro"
    if lat >= 38.5:
        # Lisboa metropolitan area vs interior Alentejo by longitude.
        return "lisboa" if lng <= -8.5 else "alentejo"
    if lat >= 37.3:
        return "alentejo"
    return "algarve"

# Portugal bounding box: continent + Açores + Madeira.
LAT_MIN, LAT_MAX = 32.0, 43.0
LNG_MIN, LNG_MAX = -32.0, -6.0

COORD_RE = re.compile(
    r"(-?\d{1,2}[.,]\d{2,8})\s*[,;\s]\s*(-?\d{1,2}[.,]\d{2,8})"
)

SECTION_HEADER_PREFIXES = ("🟦", "🟧", "🟫", "🟪", "🟩", "🟨", "🔵", "🟥", "⬛", "⬜")

_WS_RE = re.compile(r"\s+")
_PUNCT_RE = re.compile(r"[^\w\s]", re.UNICODE)


def normalise_name(name: Optional[str]) -> str:
    if not name:
        return ""
    nfkd = unicodedata.normalize("NFKD", str(name))
    no_accents = "".join(c for c in nfkd if not unicodedata.combining(c))
    no_punct = _PUNCT_RE.sub(" ", no_accents.lower())
    return _WS_RE.sub(" ", no_punct).strip()


def normalise_region(region: Optional[str]) -> str:
    if not region:
        return ""
    nfkd = unicodedata.normalize("NFKD", str(region))
    no_accents = "".join(c for c in nfkd if not unicodedata.combining(c))
    cleaned = no_accents.lower().strip()
    return REGION_MAP.get(cleaned, cleaned)


def _clean(val: Any) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ("nan", "n/a", "none"):
        return None
    return s


def _is_section_header(row: tuple) -> bool:
    """Section header rows have a single non-empty cell starting with a coloured square."""
    non_empty = [c for c in row if c is not None and str(c).strip()]
    if len(non_empty) != 1:
        return False
    first = str(non_empty[0]).strip()
    return first.startswith(SECTION_HEADER_PREFIXES) or first.startswith("  🟦") or first.startswith("  🟧") \
        or first.startswith("  🟫") or first.startswith("  🟪") or first.startswith("  🟩") \
        or first.startswith("  🔵") or first.startswith("  🟨") or first.startswith("  🟥")


_REGION_KEYWORDS = (
    # Order matters: more specific tokens first so they win over loose matches.
    ("arquipelago dos acores", "acores"),
    ("arquipelago da madeira", "madeira"),
    ("vale do tejo", "lisboa"),
    ("estremadura e ribatejo", "lisboa"),
    ("trás-os-montes", "norte"),
    ("tras-os-montes", "norte"),
    ("beira litoral", "centro"),
    ("beira interior", "centro"),
    ("alto alentejo", "alentejo"),
    ("baixo alentejo", "alentejo"),
    ("minho", "norte"),
    ("douro", "norte"),
    ("beira", "centro"),
    ("estremadura", "lisboa"),
    ("ribatejo", "lisboa"),
    ("alentejo", "alentejo"),
    ("algarve", "algarve"),
    ("acores", "acores"),
    ("açores", "acores"),
    ("madeira", "madeira"),
    ("lisboa", "lisboa"),
    ("centro", "centro"),
    ("norte", "norte"),
    ("portugal continental", "portugal"),
)


def _region_from_section_header(row: tuple) -> Optional[str]:
    """Pull the region name out of a section-header row like '🟦  NORTE  ·  40 entradas'."""
    non_empty = [c for c in row if c is not None and str(c).strip()]
    if len(non_empty) != 1:
        return None
    text = str(non_empty[0]).strip()
    nfkd = unicodedata.normalize("NFKD", text)
    no_acc = "".join(c for c in nfkd if not unicodedata.combining(c)).lower()
    for token, mapped in _REGION_KEYWORDS:
        if token in no_acc:
            return mapped
    return None


def parse_coords(text: Any) -> Optional[tuple[float, float]]:
    """Pull a (lat, lng) pair out of free-form text. Returns None if invalid or out of range."""
    if text is None:
        return None
    s = str(text).strip()
    if not s:
        return None
    m = COORD_RE.search(s)
    if not m:
        return None
    try:
        a = float(m.group(1).replace(",", "."))
        b = float(m.group(2).replace(",", "."))
    except ValueError:
        return None
    # Try (a=lat, b=lng) first
    if LAT_MIN <= a <= LAT_MAX and LNG_MIN <= b <= LNG_MAX:
        return (a, b)
    # Then swapped
    if LAT_MIN <= b <= LAT_MAX and LNG_MIN <= a <= LNG_MAX:
        return (b, a)
    return None


def _find_header_row(rows: list[tuple]) -> int:
    """Locate the header row inside the first ~6 rows.

    A header row has >= 3 non-empty cells AND mentions both a name-like word
    and a coordinate/region-like word. Sheet title rows only fill column 0.
    """
    keywords = (
        "nome", "name", "espécie", "especie", "percurso", "praia",
        "miradouro", "castelo", "palácio", "mercado", "museu", "spot",
    )
    for i in range(min(8, len(rows))):
        row = rows[i]
        if not row:
            continue
        cells = [str(c).lower() if c is not None else "" for c in row]
        non_empty = sum(1 for c in cells if c.strip())
        if non_empty < 3:
            continue
        # Strict per-cell match so the sheet-title row (which is one long cell
        # containing both "percursos" and "região") is not mistaken for a header.
        has_name = any(any(k == c.strip() or (k in c and len(c) < 40) for k in keywords) for c in cells)
        has_region_or_gps = any(
            ("região" in c or "regiao" in c or "gps" in c or "coorden" in c) and len(c) < 40
            for c in cells
        )
        if has_name and has_region_or_gps:
            return i
    return 2  # safe default for v19 format


def _build_column_map(header: list[str]) -> dict[str, int]:
    """Map semantic field → column index, based on header text."""
    cmap: dict[str, int] = {}
    for j, raw in enumerate(header):
        col = (raw or "").lower().strip()
        if not col:
            continue
        # Name (priority order: 'nome' > anything else with name-like word).
        if any(t in col for t in ("nome", "name", "espécie", "especie", "percurso", "spot")):
            cmap.setdefault("name", j)
        # "Sub-Região" must NOT match — section-header tracking gives the
        # true region for sheets that only carry sub-region columns.
        if ("região" in col or "regiao" in col) and "sub" not in col:
            cmap.setdefault("region", j)
        if "concelho" in col:
            cmap.setdefault("concelho", j)
        if "distrito" in col:
            cmap.setdefault("distrito", j)
        if "localidade" in col or "localização" in col or "localizacao" in col:
            cmap.setdefault("localidade", j)
        if "morada" in col or "endereço" in col or "endereco" in col or col == "address":
            cmap.setdefault("address", j)
        # Geocoding anchors — observation / emblematic place columns used by
        # the nature & gastronomy sheets that carry NO plain Localidade column
        # (Flora "Onde Observar", Fauna "Habitat"/"Local", Pratos/Doçaria
        # "Local Emblemático", Sopas "Restaurante Sugerido"). Mapping them to
        # `localidade` lets the geocoder resolve a precise point instead of
        # falling back to the region centroid. setdefault keeps a real
        # Localidade/Localização column (handled above) as the priority.
        if (
            "onde observar" in col
            or "habitat" in col
            or "local emblemático" in col
            or "local emblematico" in col
            or "restaurante sug" in col
            or col == "local"
        ):
            cmap.setdefault("localidade", j)
        if "descrição" in col or "descricao" in col or "description" in col:
            cmap.setdefault("description", j)
        if col == "latitude" or col.endswith(" latitude"):
            cmap.setdefault("lat_col", j)
        if col == "longitude" or col.endswith(" longitude"):
            cmap.setdefault("lng_col", j)
        if ("gps" in col or "coorden" in col) and "estacionamento" not in col:
            # Prefer the first GPS column; "GPS Estacionamento" is a fallback.
            cmap.setdefault("gps", j)
        if "estacionamento" in col and "gps" in col:
            cmap.setdefault("gps_alt", j)
        if "trilho no maps" in col:
            cmap.setdefault("gps", j)
        if col in ("id", "poi_id") or "poi_id" in col:
            cmap.setdefault("source_id", j)
        if "website" in col or col in ("url", "site", "web"):
            cmap.setdefault("website", j)
    return cmap


def _coords_from_row(row: tuple, cmap: dict[str, int]) -> tuple[Optional[tuple[float, float]], str]:
    """Return (coords, source) where source is one of: 'lat_lng_cols', 'gps_col', 'any_cell', or ''."""
    # 1. Separate Latitude / Longitude columns.
    if "lat_col" in cmap and "lng_col" in cmap:
        try:
            lat_raw = row[cmap["lat_col"]] if cmap["lat_col"] < len(row) else None
            lng_raw = row[cmap["lng_col"]] if cmap["lng_col"] < len(row) else None
            if lat_raw is not None and lng_raw is not None:
                lat = float(str(lat_raw).replace(",", "."))
                lng = float(str(lng_raw).replace(",", "."))
                if LAT_MIN <= lat <= LAT_MAX and LNG_MIN <= lng <= LNG_MAX:
                    return (lat, lng), "lat_lng_cols"
        except (ValueError, TypeError, IndexError):
            pass

    # 2. Dedicated GPS column.
    for key in ("gps", "gps_alt"):
        if key in cmap and cmap[key] < len(row):
            coords = parse_coords(row[cmap[key]])
            if coords:
                return coords, "gps_col"

    # 3. Any cell in the row (defensive — Excel data isn't always tidy).
    for cell in row:
        coords = parse_coords(cell)
        if coords:
            return coords, "any_cell"

    return None, ""


def extract_sheet(sn: str, ws) -> tuple[list[dict], dict]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4:
        return [], {"reason": "too_few_rows"}

    h_idx = _find_header_row(rows)
    header = [str(c) if c is not None else "" for c in rows[h_idx]]
    cmap = _build_column_map(header)

    if "name" not in cmap:
        # Fall back to the first non-numeric column.
        for j, h in enumerate(header):
            if h and not h.startswith("#"):
                cmap["name"] = j
                break

    category = SHEET_TO_CATEGORY.get(sn, "outros")
    pois: list[dict] = []
    stats = {
        "total": 0,
        "with_coords": 0,
        "missing_coords": 0,
        "skipped_section_header": 0,
        "skipped_empty": 0,
        "header_idx": h_idx,
        "columns": cmap,
    }

    current_region: Optional[str] = None

    for r in rows[h_idx + 1:]:
        if r is None:
            continue
        if _is_section_header(r):
            section_region = _region_from_section_header(r)
            if section_region:
                current_region = section_region
            stats["skipped_section_header"] += 1
            continue
        non_empty = sum(1 for c in r if c is not None and str(c).strip())
        if non_empty < 2:
            stats["skipped_empty"] += 1
            continue

        name = _clean(r[cmap["name"]]) if cmap.get("name") is not None and cmap["name"] < len(r) else None
        if not name:
            stats["skipped_empty"] += 1
            continue
        if name.startswith(("  🟦", "  🟧", "  🟫", "  🟪", "  🟩", "  🔵", "  🟨", "  🟥")):
            stats["skipped_section_header"] += 1
            continue
        # Header-echo rows ("Nome", "Espécie", etc.).
        if normalise_name(name) in ("nome", "name", "especie", "espécie", "id"):
            stats["skipped_empty"] += 1
            continue

        stats["total"] += 1

        coords, coord_source = _coords_from_row(r, cmap)

        region_raw = (
            _clean(r[cmap["region"]]) if "region" in cmap and cmap["region"] < len(r) else None
        )
        # Resolution order for region:
        #   1. Explicit column value (after normalisation).
        #   2. Running section-header region.
        #   3. Inference from coordinates (when coords are present).
        normalised = normalise_region(region_raw or "")
        valid_regions = {"norte", "centro", "lisboa", "alentejo", "algarve", "acores", "madeira"}
        if normalised not in valid_regions:
            if current_region and current_region in valid_regions:
                region_raw = current_region
                normalised = current_region
        if normalised not in valid_regions and coords:
            inferred = infer_region_from_coords(coords[0], coords[1])
            if inferred in valid_regions:
                region_raw = inferred
                normalised = inferred
        concelho = _clean(r[cmap["concelho"]]) if "concelho" in cmap and cmap["concelho"] < len(r) else None
        distrito = _clean(r[cmap["distrito"]]) if "distrito" in cmap and cmap["distrito"] < len(r) else None
        localidade = _clean(r[cmap["localidade"]]) if "localidade" in cmap and cmap["localidade"] < len(r) else None
        address = _clean(r[cmap["address"]]) if "address" in cmap and cmap["address"] < len(r) else None
        source_id = _clean(r[cmap["source_id"]]) if "source_id" in cmap and cmap["source_id"] < len(r) else None

        poi = {
            "name": name,
            "name_normalised": normalise_name(name),
            "sheet": sn,
            "category": category,
            "region": normalise_region(region_raw or ""),
            "region_original": region_raw,
            "concelho": concelho,
            "distrito": distrito,
            "localidade": localidade,
            "address": address,
            "source_id": source_id,
            "location": {"lat": coords[0], "lng": coords[1]} if coords else None,
            "coord_source": coord_source if coords else None,
            "coord_precision": "precise" if coords else None,
        }

        if coords:
            stats["with_coords"] += 1
        else:
            stats["missing_coords"] += 1

        pois.append(poi)

    return pois, stats


def run(excel_path: Path, out_path: Path, missing_out: Path) -> dict:
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel not found: {excel_path}")

    log.info(f"📖 Reading {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    all_pois: list[dict] = []
    per_sheet_stats: dict[str, dict] = {}

    for sn in wb.sheetnames:
        if sn in SYSTEM_SHEETS:
            continue
        ws = wb[sn]
        pois, stats = extract_sheet(sn, ws)
        if not pois:
            log.info(f"  ⏭  {sn}: skipped ({stats})")
            continue
        per_sheet_stats[sn] = {k: v for k, v in stats.items() if k not in ("columns",)}
        per_sheet_stats[sn]["category"] = SHEET_TO_CATEGORY.get(sn, "outros")
        all_pois.extend(pois)
        log.info(
            f"  ✅ {sn}: {len(pois)} POI ({stats['with_coords']} c/ coords, "
            f"{stats['missing_coords']} s/ coords)"
        )

    totals = {
        "total_pois": len(all_pois),
        "with_coords": sum(1 for p in all_pois if p["location"]),
        "without_coords": sum(1 for p in all_pois if not p["location"]),
    }
    if totals["total_pois"]:
        totals["coverage_pct"] = round(100 * totals["with_coords"] / totals["total_pois"], 1)

    by_region: dict[str, int] = {}
    by_category: dict[str, int] = {}
    for p in all_pois:
        if p["location"]:
            by_region[p["region"] or "?"] = by_region.get(p["region"] or "?", 0) + 1
            by_category[p["category"]] = by_category.get(p["category"], 0) + 1

    output = {
        "version": "v19",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source_file": excel_path.name,
        "totals": totals,
        "by_sheet": per_sheet_stats,
        "by_region": dict(sorted(by_region.items(), key=lambda x: -x[1])),
        "by_category": dict(sorted(by_category.items(), key=lambda x: -x[1])),
        "pois": [p for p in all_pois if p["location"]],
    }

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(output, ensure_ascii=False, indent=2))
    log.info(f"\n💾 Wrote {len(output['pois'])} POIs to {out_path}")

    missing = [
        {k: v for k, v in p.items() if k != "location"}
        for p in all_pois if not p["location"]
    ]
    missing_payload = {
        "version": "v19",
        "generated_at": output["generated_at"],
        "total_missing": len(missing),
        "pois": missing,
    }
    missing_out.write_text(json.dumps(missing_payload, ensure_ascii=False, indent=2))
    log.info(f"💾 Wrote {len(missing)} POIs needing geocoding to {missing_out}")

    log.info("\n" + "=" * 60)
    log.info("📊 SUMMARY")
    log.info("=" * 60)
    log.info(f"  Total POIs scanned:    {totals['total_pois']}")
    log.info(f"  With valid coords:     {totals['with_coords']} ({totals.get('coverage_pct', 0)}%)")
    log.info(f"  Missing coords:        {totals['without_coords']}")
    log.info("\n  Coverage by region:")
    for r, c in output["by_region"].items():
        log.info(f"    {r:<12} {c}")
    log.info("\n  Coverage by category:")
    for c, n in output["by_category"].items():
        log.info(f"    {c:<20} {n}")

    return output


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    p.add_argument("--out", type=Path, default=DEFAULT_OUT)
    p.add_argument("--missing-out", type=Path, default=DEFAULT_MISSING_OUT)
    args = p.parse_args()

    try:
        run(args.excel, args.out, args.missing_out)
    except FileNotFoundError as e:
        log.error(str(e))
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
