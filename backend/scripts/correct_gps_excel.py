"""
correct_gps_excel.py — Auditoria e enriquecimento de coordenadas GPS do master Excel.

Lê `PortugalVivo_BaseDados_POI_v19.xlsx` e produz três artefactos:

  1. `<basename>_audit.csv`  — uma linha por POI com: sheet, name, lat, lng,
     status (ok|missing|centroid|outside_pt|low_precision), reason.
  2. `<basename>_corrected.xlsx` — copia o Excel original e, em cada folha,
     escreve coordenadas resolvidas via Nominatim numa nova coluna
     `GPS_RESOLVED` quando estavam em falta ou eram um placeholder. O
     ficheiro original NÃO é modificado.
  3. Log estruturado em stdout (`--log-level INFO` por defeito).

Por que escolhi este desenho
----------------------------
* O master Excel é o source-of-truth da equipa editorial. Acrescentar uma
  coluna nova preserva o que o editor escreveu e deixa visível a coordenada
  proposta para cada POI; eles aprovam manualmente antes de fundir a
  coluna `GPS` definitiva.
* Nominatim é gratuito mas com rate limit estrito (1 req/s). O script
  respeita-o e suporta `--max-rows` para correr em lotes.
* Geocoding falha em muitos casos (nomes ambíguos, sopas típicas ou
  músicas tradicionais sem âncora geográfica). Esses ficam marcados como
  `still-missing` no audit, para o editor decidir se devem mesmo ter GPS.

Uso
---
  python -m scripts.correct_gps_excel \
      --input PortugalVivo_BaseDados_POI_v19.xlsx \
      --output-prefix v19 \
      --user-agent "PortugalVivo/1.0 (+contacto@portugalvivo.pt)"

Argumentos relevantes:
  --dry-run         Não escreve nada — só audit.
  --max-rows N      Geocode no máximo N linhas (debug).
  --skip-geocode    Não consulta Nominatim — só classifica o estado actual.
  --rate-limit S    Segundos entre pedidos a Nominatim (default 1.1).
"""
from __future__ import annotations

import argparse
import csv
import logging
import math
import re
import sys
import time
from pathlib import Path
from typing import Iterable, Optional

import httpx
import openpyxl
from openpyxl.utils import get_column_letter

logger = logging.getLogger("correct_gps_excel")

# ─── Envelope de Portugal (continente + ilhas) ────────────────────────────
PT_LAT_MIN, PT_LAT_MAX = 32.0, 42.5
PT_LNG_MIN, PT_LNG_MAX = -31.5, -6.0

# Centroides de capitais e cidades grandes — quando uma coordenada cai
# *exactamente* num destes pontos, é placeholder e merece geocoding.
CITY_CENTROIDS: list[tuple[str, float, float]] = [
    ("Lisboa-Marquês", 38.7223, -9.1393),
    ("Lisboa-Centro", 38.7167, -9.1333),
    ("Lisboa-Praça", 38.7139, -9.1394),
    ("Porto", 41.1496, -8.6109),
    ("Porto-Centro", 41.1579, -8.6291),
    ("Coimbra", 40.2033, -8.4103),
    ("Évora", 38.572, -7.9087),
    ("Faro", 37.0194, -7.9308),
    ("Funchal", 32.6669, -16.9241),
    ("Ponta Delgada", 37.7412, -25.6756),
    ("Braga", 41.551, -8.4254),
    ("Setúbal", 38.5244, -8.8882),
    ("Aveiro", 40.6405, -8.6538),
    ("Viseu", 40.6566, -7.9122),
    ("Guarda", 40.5371, -7.2664),
    ("Bragança", 41.8034, -6.749),
    ("Vila Real", 41.301, -7.7437),
    ("Beja", 38.0152, -7.8632),
    ("Castelo Branco", 39.8222, -7.4912),
    ("Portalegre", 39.2967, -7.4279),
    ("Leiria", 39.7437, -8.8071),
    ("Santarém", 39.2369, -8.6868),
]

# Padrões para extrair GPS de células do Excel (hyperlink, ?q=, texto livre)
RX_HYPER_AT = re.compile(r"@(-?\d+\.?\d*),(-?\d+\.?\d*)")
RX_HYPER_Q = re.compile(r"[?&]q=(-?\d+\.?\d*)[,%20]+(-?\d+\.?\d*)")
RX_TEXT = re.compile(r"(-?\d{1,2}\.\d+)\s*[,;]\s*(-?\d{1,2}\.\d+)")

# Cabeçalhos visuais que NÃO são POIs — não devem ser geocoded.
# We deliberately keep this list small: any cell that *starts* with an emoji
# is much more likely to be a section banner than a real POI name.
HEADER_MARKERS = ("🟦", "🟩", "🟨", "🟧", "🟥", "🟪", "🔵", "🟠", "🟢", "🟡", "⚪", "⬛")
HEADER_KEYWORDS = (
    "MERCADOS E FEIRAS DE PORTUGAL", "MIRADOUROS DE PORTUGAL",
    "RESTAURANTES E GASTRONOMIA", "PATRIMÓNIO FERROVIÁRIO",
    "PRAIAS COM BANDEIRA AZUL", "TERMAS E BALNEÁRIOS",
    "FARÓIS DE PORTUGAL", "FAUNA AUTÓCTONE", "FLORA AUTÓCTONE",
    "MÚSICA TRADICIONAL", "ROTAS TEMÁTICAS", "ALOJAMENTOS RURAIS",
    "FESTAS E ROMARIAS", "PERCURSOS PEDESTRES", "ECOVIAS E PASSADIÇOS",
    "PRATOS TÍPICOS", "DOÇARIA REGIONAL", "AVENTURA E NATUREZA",
    "MOINHOS E AZENHAS", "PARQUES DE CAMPISMO", "BARRAGENS E ALBUFEIRAS",
    "CASCATAS E POÇOS NATURAIS",
    "PRÉMIO PORTUGAL VIVO", "ENTRADAS", "ENTRADA",
)
# Any cell value that starts with an emoji codepoint is treated as a banner.
RX_LEADING_EMOJI = re.compile(r"^[\U0001F000-\U0001FFFF☀-➿]")
SHEET_SKIP = {"🔧 MongoDB Schema", "🔧 Ingestão Config", "📊 Dashboard",
              "Categorias", "Base de Dados POI", "🗺 Guia do Viajante",
              "🚌 Transportes"}


def looks_like_header(text: str) -> bool:
    if not text:
        return False
    t = text.strip()
    if not t:
        return False
    # Anything that starts with an emoji or is mostly emoji+spaces is a banner.
    if RX_LEADING_EMOJI.match(t):
        return True
    if any(m in t for m in HEADER_MARKERS):
        return True
    # Section titles like "EVENTOS · 240 entradas" or "AGENDA — 56 locais".
    if re.search(r"\b\d+\s+(entradas?|locais?|percursos?|registos?)\b", t, re.IGNORECASE):
        return True
    upper = t.upper()
    return any(k in upper for k in HEADER_KEYWORDS)


def extract_coords(value: object) -> Optional[tuple[float, float]]:
    """Try to pull a (lat, lng) pair out of a single cell value."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    for rx in (RX_HYPER_AT, RX_HYPER_Q, RX_TEXT):
        m = rx.search(s)
        if not m:
            continue
        try:
            lat, lng = float(m.group(1)), float(m.group(2))
        except ValueError:
            continue
        if -90 <= lat <= 90 and -180 <= lng <= 180:
            return (lat, lng)
    return None


def in_pt_envelope(lat: float, lng: float) -> bool:
    return PT_LAT_MIN <= lat <= PT_LAT_MAX and PT_LNG_MIN <= lng <= PT_LNG_MAX


def is_city_centroid(lat: float, lng: float, tol: float = 1e-4) -> Optional[str]:
    for name, clat, clng in CITY_CENTROIDS:
        if abs(lat - clat) < tol and abs(lng - clng) < tol:
            return name
    return None


def decimals(v: float) -> int:
    s = f"{v:.10f}".rstrip("0").rstrip(".")
    return len(s.split(".")[1]) if "." in s else 0


# ─── Heurística para extrair "nome do POI" e "concelho" de uma linha ─────

def best_name(cells: list[str]) -> str:
    """First plausibly-textual cell in the row.

    Skip headers, numeric labels and anything longer than 80 chars (real POI
    names are short — long strings are descriptions or marketing copy).
    """
    for c in cells[:6]:
        if not c or c.isdigit() or len(c) < 3 or len(c) > 80:
            continue
        if looks_like_header(c):
            continue
        return c.strip()
    return ""


def best_locality(cells: list[str]) -> str:
    """Try to pick up a concelho / freguesia mentioned in the row."""
    candidates = []
    for c in cells:
        if not c:
            continue
        s = c.strip()
        if 3 <= len(s) <= 50 and not s.startswith("http") and not s.isdigit():
            candidates.append(s)
    # Prefer entries that look like Title Case place names.
    return next(
        (c for c in candidates if c[:1].isupper() and " " not in c[:3] and "," not in c),
        candidates[-1] if candidates else "",
    )


# ─── Cliente Nominatim ───────────────────────────────────────────────────

class Nominatim:
    """Light Nominatim client with strict rate limiting."""

    def __init__(self, user_agent: str, rate_limit_s: float = 1.1):
        if "@" not in user_agent:
            raise ValueError("Nominatim requires a contactable user-agent (must contain '@')")
        self._user_agent = user_agent
        self._client = httpx.Client(
            base_url="https://nominatim.openstreetmap.org",
            timeout=15.0,
        )
        self._rate = rate_limit_s
        self._last_call = 0.0

    def _throttle(self) -> None:
        delta = time.monotonic() - self._last_call
        if delta < self._rate:
            time.sleep(self._rate - delta)
        self._last_call = time.monotonic()

    def search(self, query: str) -> Optional[tuple[float, float, str]]:
        if not query:
            return None
        self._throttle()
        try:
            r = self._client.get(
                "/search",
                params={
                    "q": query,
                    "format": "json",
                    "limit": 1,
                    "countrycodes": "pt",
                    "addressdetails": 0,
                },
                headers={
                    "User-Agent": self._user_agent,
                    "Accept-Language": "pt-PT,pt;q=0.8",
                },
            )
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            logger.warning("nominatim error for %r: %s", query, e)
            return None
        if not data:
            return None
        first = data[0]
        try:
            return float(first["lat"]), float(first["lon"]), first.get("display_name", "")
        except (KeyError, ValueError):
            return None

    def close(self) -> None:
        self._client.close()


# ─── Classificação por linha ─────────────────────────────────────────────

def classify(lat: Optional[float], lng: Optional[float]) -> tuple[str, str]:
    if lat is None or lng is None:
        return "missing", "no coordinates extracted"
    if not in_pt_envelope(lat, lng):
        return "outside_pt", f"out of envelope ({lat:.4f}, {lng:.4f})"
    centroid = is_city_centroid(lat, lng)
    if centroid:
        return "centroid", f"matches {centroid}"
    min_decimals = min(decimals(lat), decimals(lng))
    if min_decimals <= 2:
        return "low_precision", f"{min_decimals} decimals (≥1.1 km error)"
    return "ok", f"{min_decimals} decimals"


# ─── Main loop ───────────────────────────────────────────────────────────

def process_workbook(
    input_path: Path,
    output_prefix: str,
    user_agent: Optional[str],
    rate_limit: float,
    max_rows: Optional[int],
    skip_geocode: bool,
    dry_run: bool,
) -> None:
    logger.info("Loading %s", input_path)
    wb = openpyxl.load_workbook(input_path, data_only=False)

    audit_rows: list[dict] = []
    geocoded_pois: dict[tuple[str, int], tuple[float, float, str]] = {}
    geocode_attempts = 0
    geocode_hits = 0

    nominatim = None
    if not skip_geocode:
        if not user_agent:
            raise SystemExit("--user-agent is required for geocoding (Nominatim policy)")
        nominatim = Nominatim(user_agent, rate_limit)

    try:
        for sheet_name in wb.sheetnames:
            if sheet_name in SHEET_SKIP or sheet_name.startswith(("🔧", "📊", "🗺", "🚌")):
                continue
            ws = wb[sheet_name]
            logger.info("sheet: %s (%d rows)", sheet_name, ws.max_row)

            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                cells = [("" if c is None else str(c).strip()) for c in row]
                if not any(cells):
                    continue
                # Skip section dividers / headers
                if looks_like_header(" | ".join(cells[:5])):
                    continue
                name = best_name(cells)
                if not name:
                    continue

                # Extract any existing GPS in any cell of the row
                lat, lng = None, None
                for c in cells:
                    coords = extract_coords(c)
                    if coords:
                        lat, lng = coords
                        break

                status, reason = classify(lat, lng)
                resolved_lat: Optional[float] = None
                resolved_lng: Optional[float] = None
                resolved_label: Optional[str] = None

                if status != "ok" and not skip_geocode:
                    if max_rows is not None and geocode_attempts >= max_rows:
                        logger.info("max-rows reached, skipping further geocoding")
                    else:
                        locality = best_locality(cells)
                        query_parts = [name]
                        if locality and locality.lower() not in name.lower():
                            query_parts.append(locality)
                        query_parts.append("Portugal")
                        query = ", ".join(query_parts)
                        geocode_attempts += 1
                        logger.info("geocode[%d]: %s", geocode_attempts, query)
                        result = nominatim.search(query) if nominatim else None
                        if result:
                            geocode_hits += 1
                            resolved_lat, resolved_lng, resolved_label = result
                            geocoded_pois[(sheet_name, row_idx)] = result
                            if not in_pt_envelope(resolved_lat, resolved_lng):
                                resolved_lat = resolved_lng = None
                                resolved_label = None

                audit_rows.append({
                    "sheet": sheet_name,
                    "row": row_idx,
                    "name": name[:120],
                    "current_lat": lat if lat is not None else "",
                    "current_lng": lng if lng is not None else "",
                    "status": status,
                    "reason": reason,
                    "resolved_lat": resolved_lat if resolved_lat is not None else "",
                    "resolved_lng": resolved_lng if resolved_lng is not None else "",
                    "resolved_source": resolved_label or "",
                })
    finally:
        if nominatim:
            nominatim.close()

    # ── Write artefacts ─────────────────────────────────────────────────
    audit_path = Path(f"{output_prefix}_audit.csv")
    with audit_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(audit_rows[0].keys()) if audit_rows else [])
        writer.writeheader()
        writer.writerows(audit_rows)
    logger.info("wrote %s (%d rows)", audit_path, len(audit_rows))

    by_status: dict[str, int] = {}
    for r in audit_rows:
        by_status[r["status"]] = by_status.get(r["status"], 0) + 1
    logger.info("status summary: %s", by_status)
    logger.info("geocode: %d attempts, %d hits (%.1f%%)", geocode_attempts, geocode_hits,
                100 * geocode_hits / max(1, geocode_attempts))

    if dry_run:
        logger.info("dry-run: not writing corrected workbook")
        return

    corrected_path = Path(f"{output_prefix}_corrected.xlsx")
    wb2 = openpyxl.load_workbook(input_path, data_only=False)
    for sheet_name in wb2.sheetnames:
        if sheet_name in SHEET_SKIP or sheet_name.startswith(("🔧", "📊", "🗺", "🚌")):
            continue
        ws = wb2[sheet_name]
        # Add a single new column to the right with the resolved coordinates.
        new_col = ws.max_column + 1
        ws.cell(row=1, column=new_col, value="GPS_RESOLVED")
        for (sn, row_idx), (lat, lng, label) in geocoded_pois.items():
            if sn != sheet_name:
                continue
            ws.cell(row=row_idx, column=new_col,
                    value=f"{lat:.6f}, {lng:.6f}  # {label[:80]}")
    wb2.save(corrected_path)
    logger.info("wrote %s", corrected_path)


def main(argv: Optional[Iterable[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output-prefix", default="poi_gps")
    parser.add_argument("--user-agent",
                        help="Contact e-mail in the Nominatim User-Agent. Required unless --skip-geocode.")
    parser.add_argument("--rate-limit", type=float, default=1.1,
                        help="Seconds between Nominatim requests (default 1.1)")
    parser.add_argument("--max-rows", type=int, default=None,
                        help="Stop geocoding after this many requests (debug)")
    parser.add_argument("--skip-geocode", action="store_true",
                        help="Audit only — do not call Nominatim")
    parser.add_argument("--dry-run", action="store_true",
                        help="Audit only — do not write the corrected workbook")
    parser.add_argument("--log-level", default="INFO")
    args = parser.parse_args(list(argv) if argv is not None else None)

    logging.basicConfig(
        level=getattr(logging, args.log_level.upper(), logging.INFO),
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    if not args.input.is_file():
        parser.error(f"input not found: {args.input}")

    process_workbook(
        input_path=args.input,
        output_prefix=args.output_prefix,
        user_agent=args.user_agent,
        rate_limit=args.rate_limit,
        max_rows=args.max_rows,
        skip_geocode=args.skip_geocode,
        dry_run=args.dry_run,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
