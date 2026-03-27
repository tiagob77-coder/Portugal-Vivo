"""
Import Excel v19 — Portugal Vivo
Importa TODOS os POIs do ficheiro PortugalVivo_BaseDados_POI_v19.xlsx
para a colecção heritage_items no MongoDB.

Suporta 43+ folhas com estruturas diferentes.
Extrai GPS de HYPERLINK formulas.
Mapeia categorias para as 44 subcategorias do schema.
"""

import asyncio
import re
import uuid
import os
import logging
from pathlib import Path
from datetime import datetime, timezone
from typing import Optional, Dict, List, Any, Tuple

import openpyxl
from motor.motor_asyncio import AsyncIOMotorClient
from dotenv import load_dotenv

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ─── Sheet → Category/Subcategory Mapping ──────────────────────────────

SHEET_MAP: Dict[str, Dict[str, Any]] = {
    # Natureza
    "Percursos Pedestres": {
        "category": "percursos_pedestres",
        "main_category": "territorio_natureza",
        "main_category_name": "Natureza",
        "category_name": "Percursos Pedestres",
        "icon": "hiking",
        "columns": {"name": 1, "sub_region": 2, "distance": 3, "duration": 4, "difficulty": 5, "type": 6, "altitude": 7, "best_season": 8, "pois_desc": 9, "gps": 10, "address": 11},
    },
    "Ecovias e Passadiços": {
        "category": "ecovias_passadicos",
        "main_category": "territorio_natureza",
        "main_category_name": "Natureza",
        "category_name": "Ecovias e Passadiços",
        "icon": "directions-walk",
        "columns": {"name": 1, "sub_region": 2, "gps_start": 3, "gps_end": 4, "distance": 5, "type": 6, "services": 7, "description": 8, "poi_main": 9, "gps": 10},
    },
    "Aventura e Natureza": {
        "category": "aventura_natureza",
        "main_category": "territorio_natureza",
        "main_category_name": "Natureza",
        "category_name": "Aventura e Natureza",
        "icon": "terrain",
        "columns": {"name": 1, "type": 2, "region": 3, "address": 4, "description": 5, "gps": -1},
    },
    "Natureza Especializada": {
        "category": "natureza_especializada",
        "main_category": "territorio_natureza",
        "main_category_name": "Natureza",
        "category_name": "Natureza Especializada",
        "icon": "eco",
        "columns": {"name": 1, "type": 2, "region": 3, "address": 4, "description": 5, "gps": -1},
    },
    "Fauna Autóctone": {
        "category": "fauna_autoctone",
        "main_category": "territorio_natureza",
        "main_category_name": "Natureza",
        "category_name": "Fauna Autóctone",
        "icon": "pets",
        "columns": {"name": 1, "type": 2, "region": 3, "habitat": 4, "location": 5, "rarity": 6, "description": 7, "gps": 8},
    },
    "Flora Autóctone": {
        "category": "flora_autoctone",
        "main_category": "territorio_natureza",
        "main_category_name": "Natureza",
        "category_name": "Flora Autóctone",
        "icon": "local-florist",
        "columns": {"name": 1, "status": 2, "region": 3, "flowering": 4, "curiosity": 5, "where_to_see": 6, "description": 7, "gps": 8},
    },
    "Flora Botânica": {
        "category": "flora_botanica",
        "main_category": "territorio_natureza",
        "main_category_name": "Natureza",
        "category_name": "Flora Botânica",
        "icon": "park",
        "columns": {"name": 1, "status": 2, "region": 3, "flowering": 4, "curiosity": 5, "where_to_see": 6, "description": 7, "gps": 8},
    },
    "Biodiversidade | Avistamentos": {
        "category": "biodiversidade_avistamentos",
        "main_category": "territorio_natureza",
        "main_category_name": "Natureza",
        "category_name": "Biodiversidade",
        "icon": "visibility",
        "columns": {"name": 1, "type": 2, "region": 3, "description": 4, "gps": -1},
    },
    "Miradouros Portugal": {
        "category": "miradouros",
        "main_category": "territorio_natureza",
        "main_category_name": "Natureza",
        "category_name": "Miradouros",
        "icon": "panorama",
        "columns": {"name": 1, "region": 2, "location": 3, "walk": 4, "sun": 5, "btt": 6, "description": 7, "gps": 8},
    },
    "Barragens e Albufeiras": {
        "category": "barragens_albufeiras",
        "main_category": "territorio_natureza",
        "main_category_name": "Natureza",
        "category_name": "Barragens e Albufeiras",
        "icon": "water",
        "columns": {"name": 1, "region": 2, "address": 3, "description": 4, "gps": -1},
    },
    "Cascatas e Poços Naturais": {
        "category": "cascatas_pocos",
        "main_category": "territorio_natureza",
        "main_category_name": "Natureza",
        "category_name": "Cascatas e Poços Naturais",
        "icon": "waves",
        "columns": {"name": 1, "region": 2, "address": 3, "description": 4, "gps": -1},
    },
    "Praias Fluviais": {
        "category": "praias_fluviais",
        "main_category": "territorio_natureza",
        "main_category_name": "Natureza",
        "category_name": "Praias Fluviais",
        "icon": "pool",
        "columns": {"name": 1, "region": 2, "address": 3, "type": 4, "quality": 5, "description": 6, "gps": -1},
    },
    "Arqueologia, Geologia e Mineral": {
        "category": "arqueologia_geologia",
        "main_category": "territorio_natureza",
        "main_category_name": "Natureza",
        "category_name": "Arqueologia e Geologia",
        "icon": "diamond",
        "columns": {"name": 1, "type": 2, "region": 3, "address": 4, "description": 5, "gps": -1},
    },
    "Moinhos e Azenhas": {
        "category": "moinhos_azenhas",
        "main_category": "territorio_natureza",
        "main_category_name": "Natureza",
        "category_name": "Moinhos e Azenhas",
        "icon": "settings",
        "columns": {"name": 1, "type": 2, "region": 3, "address": 4, "description": 5, "gps": -1},
    },
    # História & Património
    "Castelos": {
        "category": "castelos",
        "main_category": "historia_patrimonio",
        "main_category_name": "História & Património",
        "category_name": "Castelos",
        "icon": "fort",
        "columns": {"name": 1, "period": 2, "region": 3, "address": 4, "schedule": 5, "price": 6, "website": 7, "description": 8, "gps": 9},
    },
    "Palácios e Solares": {
        "category": "palacios_solares",
        "main_category": "historia_patrimonio",
        "main_category_name": "História & Património",
        "category_name": "Palácios e Solares",
        "icon": "villa",
        "columns": {"name": 1, "period": 2, "region": 3, "address": 4, "schedule": 5, "price": 6, "website": 7, "description": 8, "gps": 9},
    },
    "Museus": {
        "category": "museus",
        "main_category": "historia_patrimonio",
        "main_category_name": "História & Património",
        "category_name": "Museus",
        "icon": "museum",
        "columns": {"name": 1, "region": 2, "address": 3, "postal_code": 4, "cat": 5, "typology": 6, "schedule": 7, "website": 8, "gps": 9, "description": 10},
    },
    "Ofícios e Artesanato": {
        "category": "oficios_artesanato",
        "main_category": "historia_patrimonio",
        "main_category_name": "História & Património",
        "category_name": "Ofícios e Artesanato",
        "icon": "handyman",
        "columns": {"name": 1, "type": 2, "region": 3, "address": 4, "description": 5, "gps": -1},
    },
    "Termas e Banhos": {
        "category": "termas_banhos",
        "main_category": "historia_patrimonio",
        "main_category_name": "História & Património",
        "category_name": "Termas e Banhos",
        "icon": "hot-tub",
        "columns": {"name": 1, "region": 2, "address": 3, "description": 4, "gps": -1},
    },
    "Património Ferroviário": {
        "category": "patrimonio_ferroviario",
        "main_category": "historia_patrimonio",
        "main_category_name": "História & Património",
        "category_name": "Património Ferroviário",
        "icon": "train",
        "columns": {"name": 1, "type": 2, "region": 3, "address": 4, "description": 5, "gps": -1},
    },
    "Arte Urbana e Intervenção": {
        "category": "arte_urbana",
        "main_category": "historia_patrimonio",
        "main_category_name": "História & Património",
        "category_name": "Arte Urbana",
        "icon": "palette",
        "columns": {"name": 1, "type": 2, "region": 3, "address": 4, "description": 5, "gps": -1},
    },
    # Gastronomia
    "Restaurantes e Gastronomia": {
        "category": "restaurantes_gastronomia",
        "main_category": "gastronomia_produtos",
        "main_category_name": "Gastronomia",
        "category_name": "Restaurantes",
        "icon": "restaurant",
        "columns": {"name": 1, "typology": 2, "region": 3, "address": 4, "contact": 5, "price": 6, "location": 7, "highlight": 8, "description": 9, "gps": 10},
    },
    "Tabernas Históricas": {
        "category": "tabernas_historicas",
        "main_category": "gastronomia_produtos",
        "main_category_name": "Gastronomia",
        "category_name": "Tabernas Históricas",
        "icon": "local-bar",
        "columns": {"name": 1, "address": 2, "city": 3, "highlight": 4, "website": 5, "description": 6, "contact": 7, "price": 8, "gps": 9},
    },
    "Mercados e Feiras": {
        "category": "mercados_feiras",
        "main_category": "gastronomia_produtos",
        "main_category_name": "Gastronomia",
        "category_name": "Mercados e Feiras",
        "icon": "storefront",
        "columns": {"name": 1, "type": 2, "region": 3, "address": 4, "description": 5, "gps": -1},
    },
    "Produtores DOP e Locais": {
        "category": "produtores_dop",
        "main_category": "gastronomia_produtos",
        "main_category_name": "Gastronomia",
        "category_name": "Produtores DOP",
        "icon": "agriculture",
        "columns": {"name": 1, "region": 2, "address": 3, "product": 4, "description": 5, "gps": -1},
    },
    "Agroturismo e Enoturismo": {
        "category": "agroturismo_enoturismo",
        "main_category": "gastronomia_produtos",
        "main_category_name": "Gastronomia",
        "category_name": "Agroturismo e Enoturismo",
        "icon": "wine-bar",
        "columns": {"name": 1, "region": 2, "address": 3, "product": 4, "description": 5, "website": 6, "contact": 7, "gps": 8},
    },
    "Pratos Típicos": {
        "category": "pratos_tipicos",
        "main_category": "gastronomia_produtos",
        "main_category_name": "Gastronomia",
        "category_name": "Pratos Típicos",
        "icon": "lunch-dining",
        "columns": {"name": 1, "region": 2, "description": 3, "gps": -1},
    },
    "Sopas Típicas": {
        "category": "sopas_tipicas",
        "main_category": "gastronomia_produtos",
        "main_category_name": "Gastronomia",
        "category_name": "Sopas Típicas",
        "icon": "soup-kitchen",
        "columns": {"name": 2, "description": 3, "location1": 4, "city1": 5, "gps_raw": 6},
        "special_format": "sopas",
    },
    "Doçaria Regional": {
        "category": "docaria_regional",
        "main_category": "gastronomia_produtos",
        "main_category_name": "Gastronomia",
        "category_name": "Doçaria Regional",
        "icon": "cake",
        "columns": {"name": 1, "region": 2, "description": 3, "location1": 4, "gps1": 5, "location2": 6, "gps2": 7},
    },
    # Cultura
    "Música Tradicional": {
        "category": "musica_tradicional",
        "main_category": "cultura_viva",
        "main_category_name": "Cultura",
        "category_name": "Música Tradicional",
        "icon": "music-note",
        "columns": {"name": 1, "type": 2, "region": 3, "description": 4, "gps": -1},
    },
    "Festivais de Música": {
        "category": "festivais_musica",
        "main_category": "cultura_viva",
        "main_category_name": "Cultura",
        "category_name": "Festivais de Música",
        "icon": "celebration",
        "columns": {"name": 1, "type": 2, "region": 3, "date": 4, "description": 5, "gps": -1},
    },
    "Festas e Romarias": {
        "category": "festas_romarias",
        "main_category": "cultura_viva",
        "main_category_name": "Cultura",
        "category_name": "Festas e Romarias",
        "icon": "festival",
        "columns": {"name": 1, "city": 2, "date": 3, "rarity": 4, "description": 5, "gps": 6, "website": 7, "location": 8},
    },
    # Mar & Praias
    "Surf": {
        "category": "surf",
        "main_category": "praias_mar",
        "main_category_name": "Mar & Praias",
        "category_name": "Surf",
        "icon": "surfing",
        "columns": {"name": 1, "region": 2, "address": 3, "level": 4, "wave_type": 5, "best_season": 6, "description": 7, "gps": 8},
    },
    "Praias Bandeira Azul": {
        "category": "praias_bandeira_azul",
        "main_category": "praias_mar",
        "main_category_name": "Mar & Praias",
        "category_name": "Praias Bandeira Azul",
        "icon": "flag",
        "columns": {"name": 1, "region": 2, "city": 3, "type": 4, "island": 5, "quality": 6, "description": 7, "gps": 8},
    },
    # Experiências
    "Rotas Temáticas": {
        "category": "rotas_tematicas",
        "main_category": "experiencias_rotas",
        "main_category_name": "Experiências",
        "category_name": "Rotas Temáticas",
        "icon": "route",
        "columns": {"name": 1, "type": 2, "region": 3, "duration": 4, "transport": 5, "highlights": 6, "description": 7, "gps": 8},
    },
    "Grande Expedição 2026": {
        "category": "grande_expedicao",
        "main_category": "experiencias_rotas",
        "main_category_name": "Experiências",
        "category_name": "Grande Expedição 2026",
        "icon": "explore",
        "columns": {"name": 1, "type": 2, "region": 3, "description": 4, "gps": -1},
    },
    "Pérolas de Portugal": {
        "category": "perolas_portugal",
        "main_category": "experiencias_rotas",
        "main_category_name": "Experiências",
        "category_name": "Pérolas de Portugal",
        "icon": "diamond",
        "columns": {"name": 1, "region": 2, "description": 3, "gps": -1},
    },
    # Alojamento
    "Parques de Campismo": {
        "category": "parques_campismo",
        "main_category": "experiencias_rotas",
        "main_category_name": "Experiências",
        "category_name": "Parques de Campismo",
        "icon": "camping",
        "columns": {"name": 1, "type": 2, "region": 3, "address": 4, "description": 5, "gps": -1},
    },
    "Pousadas de Juventude": {
        "category": "pousadas_juventude",
        "main_category": "experiencias_rotas",
        "main_category_name": "Experiências",
        "category_name": "Pousadas de Juventude",
        "icon": "hotel",
        "columns": {"name": 1, "region": 2, "address": 3, "description": 4, "gps": -1},
    },
    "Alojamentos Rurais": {
        "category": "alojamentos_rurais",
        "main_category": "experiencias_rotas",
        "main_category_name": "Experiências",
        "category_name": "Alojamentos Rurais",
        "icon": "cottage",
        "columns": {"name": 1, "typology": 2, "region": 3, "address": 4, "ambience": 5, "count": 6, "reference": 7, "gps": 8, "description": 9},
    },
    # Património Técnico
    "Faróis": {
        "category": "farois",
        "main_category": "historia_patrimonio",
        "main_category_name": "História & Património",
        "category_name": "Faróis",
        "icon": "lighthouse",
        "columns": {"name": 1, "region": 2, "address": 3, "description": 4, "gps": -1},
    },
    # B2B
    "Entidades e Operadores": {
        "category": "entidades_operadores",
        "main_category": "experiencias_rotas",
        "main_category_name": "Experiências",
        "category_name": "Entidades e Operadores",
        "icon": "business",
        "columns": {"name": 1, "type": 2, "region": 3, "address": 4, "description": 5, "gps": -1},
    },
    "Agentes Turísticos": {
        "category": "agentes_turisticos",
        "main_category": "experiencias_rotas",
        "main_category_name": "Experiências",
        "category_name": "Agentes Turísticos",
        "icon": "tour",
        "columns": {"name": 1, "type": 2, "region": 3, "address": 4, "description": 5, "gps": -1},
    },
}

# ─── Region detection from separator rows ─────────────────────────────

REGION_MAP = {
    "norte": "norte",
    "centro": "centro",
    "lisboa": "lisboa",
    "alentejo": "alentejo",
    "algarve": "algarve",
    "açores": "acores",
    "acores": "acores",
    "madeira": "madeira",
}

REGION_COORDS = {
    "norte": (41.15, -8.61),
    "centro": (40.21, -8.43),
    "lisboa": (38.72, -9.14),
    "alentejo": (38.57, -7.91),
    "algarve": (37.02, -7.93),
    "acores": (37.74, -25.66),
    "madeira": (32.65, -16.91),
}

def detect_region_from_row(text: str) -> Optional[str]:
    """Detect if a row is a region separator (e.g., '🟦 NORTE · 86 percursos')"""
    if not text:
        return None
    text_lower = text.lower().strip()
    for key, region in REGION_MAP.items():
        if key in text_lower and ("·" in text or "—" in text or "entradas" in text_lower or "percursos" in text_lower or "locais" in text_lower):
            return region
    return None


def extract_gps_from_hyperlink(value: str) -> Optional[Tuple[float, float]]:
    """Extract lat/lng from a Google Maps HYPERLINK formula."""
    if not value or not isinstance(value, str):
        return None
    
    # Pattern 1: /@lat,lng,zoom
    m = re.search(r'@(-?\d+\.?\d*),(-?\d+\.?\d*)', value)
    if m:
        lat, lng = float(m.group(1)), float(m.group(2))
        if -90 <= lat <= 90 and -180 <= lng <= 180:
            return (lat, lng)
    
    # Pattern 2: ?q=lat,lng or ?q=...&ll=lat,lng
    m = re.search(r'[?&]q=(-?\d+\.?\d*)[,%20]+(-?\d+\.?\d*)', value)
    if m:
        lat, lng = float(m.group(1)), float(m.group(2))
        if -90 <= lat <= 90 and -180 <= lng <= 180:
            return (lat, lng)
    
    return None


def extract_gps_from_text(value: str) -> Optional[Tuple[float, float]]:
    """Extract GPS from raw coordinate text like '41.276400, -8.283100'"""
    if not value or not isinstance(value, str):
        return None
    m = re.search(r'(-?\d+\.?\d+)\s*[,;]\s*(-?\d+\.?\d+)', value)
    if m:
        lat, lng = float(m.group(1)), float(m.group(2))
        if -90 <= lat <= 90 and -180 <= lng <= 180:
            return (lat, lng)
    return None


def make_slug(name: str) -> str:
    """Generate URL-safe slug from name."""
    import unicodedata
    slug = unicodedata.normalize('NFKD', name).encode('ascii', 'ignore').decode()
    slug = re.sub(r'[^\w\s-]', '', slug.lower().strip())
    slug = re.sub(r'[-\s]+', '-', slug)
    return slug[:100]


def cell_str(val) -> str:
    """Convert cell value to string, handling None and numbers."""
    if val is None:
        return ""
    return str(val).strip()


def parse_sheet(ws, sheet_name: str, config: Dict) -> List[Dict]:
    """Parse a single Excel sheet into POI documents."""
    items = []
    cols = config["columns"]
    current_region = "norte"  # default
    
    special = config.get("special_format")
    
    header_row = 3  # Most sheets have headers in row 3
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1)):
        actual_row = row_idx + 1
        if actual_row <= header_row:
            continue
        
        cells = [cell.value for cell in row]
        if not cells or not any(cells):
            continue
        
        first = cell_str(cells[0])
        
        # Check for region separator
        region_detected = detect_region_from_row(first)
        if region_detected:
            current_region = region_detected
            continue
        
        # Also check second cell for region (some sheets have it there)
        if len(cells) > 1:
            region_from_second = detect_region_from_row(cell_str(cells[1]))
            if region_from_second and not first.isdigit():
                current_region = region_from_second
                continue
        
        # Skip non-data rows (no number in first column)
        if not first or (not first.isdigit() and not re.match(r'^\d+$', first.strip())):
            # But some sheets have region in col 0 without separator format
            for key, region in REGION_MAP.items():
                if key in first.lower() and len(first) < 60:
                    current_region = region
            continue
        
        # Get name
        name_col = cols.get("name", 1)
        name = cell_str(cells[name_col]) if name_col < len(cells) else ""
        if not name or len(name) < 2:
            continue
        
        # Get description - find it in different possible columns
        desc = ""
        desc_col = cols.get("description")
        if desc_col and desc_col < len(cells):
            desc = cell_str(cells[desc_col])
        if not desc:
            # Try to find description in POIs column for trails
            pois_col = cols.get("pois_desc")
            if pois_col and pois_col < len(cells):
                desc = cell_str(cells[pois_col])
        
        # Get region from data if available
        region = current_region
        region_col = cols.get("region")
        if region_col and region_col < len(cells):
            r_val = cell_str(cells[region_col]).lower()
            for key, mapped in REGION_MAP.items():
                if key in r_val:
                    region = mapped
                    break
        
        # Get address
        address = ""
        for addr_key in ["address", "city", "location", "sub_region"]:
            addr_col = cols.get(addr_key)
            if addr_col and addr_col < len(cells):
                a_val = cell_str(cells[addr_col])
                if a_val and len(a_val) > 1:
                    address = a_val
                    break
        
        # Extract GPS coordinates
        lat, lng = None, None
        gps_col = cols.get("gps", -1)
        
        # Try specific GPS column
        if gps_col >= 0 and gps_col < len(cells):
            coords = extract_gps_from_hyperlink(cell_str(cells[gps_col]))
            if coords:
                lat, lng = coords
        
        # Try last column (many sheets put GPS there)
        if lat is None:
            for ci in range(len(cells) - 1, max(0, len(cells) - 4), -1):
                v = cell_str(cells[ci])
                if "HYPERLINK" in v or "google.com/maps" in v:
                    coords = extract_gps_from_hyperlink(v)
                    if coords:
                        lat, lng = coords
                        break
        
        # Try raw GPS text columns
        if lat is None:
            gps_raw_col = cols.get("gps_raw")
            if gps_raw_col and gps_raw_col < len(cells):
                coords = extract_gps_from_text(cell_str(cells[gps_raw_col]))
                if coords:
                    lat, lng = coords
        
        if lat is None:
            for ci in range(len(cells)):
                v = cell_str(cells[ci])
                coords = extract_gps_from_text(v)
                if coords:
                    lat, lng = coords
                    break
        
        # Fallback: region center with random offset
        if lat is None:
            import random
            base = REGION_COORDS.get(region, (39.5, -8.0))
            lat = base[0] + random.uniform(-0.5, 0.5)
            lng = base[1] + random.uniform(-0.5, 0.5)
        
        # Build extra fields
        extra = {}
        for key in ["distance", "duration", "difficulty", "type", "altitude", "best_season",
                     "period", "schedule", "price", "website", "contact", "typology",
                     "highlight", "level", "wave_type", "quality", "rarity", "product",
                     "walk", "island", "ambience"]:
            col_idx = cols.get(key)
            if col_idx and col_idx < len(cells):
                val = cell_str(cells[col_idx])
                if val and len(val) > 0:
                    extra[key] = val
        
        slug = make_slug(name)
        poi_id = f"poi_{slug}_{uuid.uuid4().hex[:6]}"
        
        item = {
            "id": poi_id,
            "name": name,
            "slug": slug,
            "description": desc[:2000] if desc else f"{name} — {config['category_name']} em {region.title()}",
            "category": config["category"],
            "category_name": config["category_name"],
            "main_category": config["main_category"],
            "main_category_name": config["main_category_name"],
            "icon": config["icon"],
            "region": region,
            "address": address,
            "location": {"lat": round(lat, 6), "lng": round(lng, 6)},
            "image_url": "",
            "tags": [config["category_name"], region, config["main_category_name"]],
            "folha_origem": sheet_name,
            "estado": "publicado",
            "created_at": datetime.now(timezone.utc),
            "updated_at": datetime.now(timezone.utc),
        }
        
        if extra:
            item["metadata"] = extra
        
        items.append(item)
    
    return items


async def import_all(db, excel_path: str = None):
    """Import all sheets from the Excel file into heritage_items."""
    if excel_path is None:
        excel_path = str(ROOT_DIR / "PortugalVivo_BaseDados_POI_v19.xlsx")
    
    if not Path(excel_path).exists():
        logger.warning("Excel file not found: %s — skipping import", excel_path)
        return 0
    
    logger.info("📚 Loading Excel: %s", excel_path)
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=False)
    
    total_imported = 0
    all_items = []
    
    for sheet_name, config in SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found in Excel", sheet_name)
            continue
        
        ws = wb[sheet_name]
        items = parse_sheet(ws, sheet_name, config)
        all_items.extend(items)
        logger.info("  📄 %s: %d POIs parsed", sheet_name, len(items))
    
    wb.close()
    
    if not all_items:
        logger.warning("No items parsed from Excel")
        return 0
    
    # Clear existing heritage_items and insert new ones
    existing = await db.heritage_items.count_documents({})
    if existing > 0:
        logger.info("🗑️ Clearing %d existing heritage_items", existing)
        await db.heritage_items.delete_many({})
    
    # Batch insert
    batch_size = 500
    for i in range(0, len(all_items), batch_size):
        batch = all_items[i:i + batch_size]
        await db.heritage_items.insert_many(batch)
        logger.info("  ✅ Inserted batch %d-%d", i + 1, min(i + batch_size, len(all_items)))
    
    total_imported = len(all_items)
    
    # Create indexes
    await db.heritage_items.create_index("id", unique=True)
    await db.heritage_items.create_index("category")
    await db.heritage_items.create_index("main_category")
    await db.heritage_items.create_index("region")
    await db.heritage_items.create_index("slug")
    await db.heritage_items.create_index([("name", "text"), ("description", "text")], default_language="portuguese")
    await db.heritage_items.create_index([("location.lat", 1), ("location.lng", 1)])
    await db.heritage_items.create_index("folha_origem")
    
    logger.info("🎉 Total imported: %d POIs from %d sheets", total_imported, len(SHEET_MAP))
    
    # Print summary by category
    pipeline = [
        {"$group": {"_id": "$main_category_name", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
    ]
    summary = await db.heritage_items.aggregate(pipeline).to_list(50)
    for s in summary:
        logger.info("  📊 %s: %d", s["_id"], s["count"])
    
    return total_imported


async def main():
    mongo_url = os.environ['MONGO_URL']
    client = AsyncIOMotorClient(mongo_url)
    db = client[os.environ['DB_NAME']]
    count = await import_all(db)
    print(f"\n✅ Importados {count} POIs com sucesso!")
    client.close()


if __name__ == "__main__":
    asyncio.run(main())
