"""
Universal Importer v2 for PortugalVivo_BaseDados_POI_v19.xlsx
Reads HYPERLINK formulas to extract GPS coordinates from Google Maps URLs.
"""
import openpyxl
import re
import uuid
from datetime import datetime, timezone
from pymongo import MongoClient
import os
from dotenv import load_dotenv

load_dotenv()

MONGO_URL = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = os.environ.get("DB_NAME", "portugal_vivo")

SHEET_CATEGORY_MAP = {
    "🎵 Música Tradicional": "musica_tradicional",
    "🗺 Guia do Viajante": "guia_viajante",
    "🚌 Transportes": "transportes",
    "Sopas Típicas": "pratos_tipicos",
    "🌊 Barragens e Albufeiras": "barragens_albufeiras",
    "Pratos Típicos": "pratos_tipicos",
    "🍬 Doçaria Regional": "docaria_regional",
    "🎪 Festivais de Música": "festivais_musica",
    "Festas e Romarias": "festas_romarias",
    "Percursos Pedestres": "percursos_pedestres",
    "Ofícios e Artesanato": "oficios_artesanato",
    "Mercados e Feiras": "mercados_feiras",
    "Tabernas Históricas": "tabernas_historicas",
    "Museus": "museus",
    "Fauna Autóctone": "fauna_autoctone",
    "Flora Autóctone": "flora_autoctone",
    "Aventura e Natureza": "aventura_natureza",
    "Natureza Especializada": "natureza_especializada",
    "Flora Botânica": "flora_botanica",
    "Castelos": "castelos",
    "Palácios e Solares": "palacios_solares",
    "Surf": "surf",
    "Praias Fluviais": "praias_fluviais",
    "🧀 Produtores DOP e Locais": "produtores_dop",
    "Restaurantes e Gastronomia": "restaurantes_gastronomia",
    "Parques de Campismo": "parques_campismo",
    "Pousadas de Juventude": "pousadas_juventude",
    "Rotas Temáticas": "rotas_tematicas",
    "Grande Expedição 2026": "grande_expedicao",
    "Biodiversidade | Avistamentos": "biodiversidade_avistamentos",
    "Miradouros Portugal": "miradouros",
    "Património Ferroviário": "patrimonio_ferroviario",
    "Arte Urbana e Intervenção": "arte_urbana",
    "Cascatas e Poços Naturais": "cascatas_pocos",
    "Termas e Banhos": "termas_banhos",
    "💎 Pérolas de Portugal": "perolas_portugal",
    "Moinhos e Azenhas": "moinhos_azenhas",
    "Arqueologia, Geologia e Mineral": "arqueologia_geologia",
    "Ecovias e Passadiços": "ecovias_passadicos",
    "Praias Bandeira Azul": "praias_bandeira_azul",
    "Alojamentos Rurais": "alojamentos_rurais",
    "Entidades e Operadores": "entidades_operadores",
    "Agentes Turísticos": "agentes_turisticos",
    "Agroturismo e Enoturismo": "agroturismo_enoturismo",
    "Faróis": "natureza_especializada",
}

SKIP_SHEETS = {"📊 Dashboard", "Categorias", "Base de Dados POI"}

REGION_MAP = {
    "norte": "norte", "porto": "norte", "minho": "norte", "trás-os-montes": "norte",
    "tras-os-montes": "norte", "douro": "norte", "braga": "norte", "viana": "norte",
    "bragança": "norte", "vila real": "norte",
    "centro": "centro", "beira": "centro", "coimbra": "centro", "aveiro": "centro",
    "leiria": "centro", "viseu": "centro", "guarda": "centro", "castelo branco": "centro",
    "beira litoral": "centro", "beira alta": "centro", "beira interior": "centro",
    "beira baixa": "centro",
    "lisboa": "lisboa", "setúbal": "lisboa", "estremadura": "lisboa",
    "ribatejo": "lisboa", "santarém": "lisboa",
    "alentejo": "alentejo", "évora": "alentejo", "portalegre": "alentejo",
    "beja": "alentejo", "alto alentejo": "alentejo", "baixo alentejo": "alentejo",
    "alentejo litoral": "alentejo", "alentejo central": "alentejo",
    "algarve": "algarve", "faro": "algarve",
    "açores": "acores", "acores": "acores", "azores": "acores",
    "são miguel": "acores", "terceira": "acores", "faial": "acores",
    "madeira": "madeira", "funchal": "madeira", "porto santo": "madeira",
}


def normalize_region(region_str):
    if not region_str:
        return "norte"
    r = str(region_str).strip().lower()
    if r in REGION_MAP:
        return REGION_MAP[r]
    for key, val in REGION_MAP.items():
        if key in r or r in key:
            return val
    return "norte"


def extract_gps_from_hyperlink(formula):
    """Extract lat/lng from HYPERLINK formula with Google Maps URL."""
    if not formula:
        return None
    s = str(formula)
    # Match @lat,lng pattern in Google Maps URLs
    match = re.search(r'@([-]?\d+\.?\d*),([-]?\d+\.?\d*)', s)
    if match:
        lat, lng = float(match.group(1)), float(match.group(2))
        if -90 <= lat <= 90 and -180 <= lng <= 180:
            return {"lat": lat, "lng": lng}
    # Match q=lat,lng or place/lat,lng
    match = re.search(r'[/q=]([-]?\d+\.\d+)[,+]([-]?\d+\.\d+)', s)
    if match:
        lat, lng = float(match.group(1)), float(match.group(2))
        if -90 <= lat <= 90 and -180 <= lng <= 180:
            return {"lat": lat, "lng": lng}
    return None


def extract_gps_plain(value):
    """Extract GPS from plain text coordinates."""
    if not value:
        return None
    s = str(value).strip()
    s = re.sub(r'[📍🗺️🏰🏛👁🏄🎉🍷⛺🎒🗺🧀🚂🎨💦♨️💎⚙️🪨🚶🏖🛖🏢🧭🍇🦊🌿🦅🌸🎪🛒🪡🍽🍬🎵🚌🌊🧗]', '', s)
    # Plain coordinates
    match = re.search(r'([-]?\d+\.\d{2,})[,;\s]+([-]?\d+\.\d{2,})', s)
    if match:
        lat, lng = float(match.group(1)), float(match.group(2))
        if -90 <= lat <= 90 and -180 <= lng <= 180:
            return {"lat": lat, "lng": lng}
    return None


def find_header_row(ws, max_rows=10):
    for i in range(1, max_rows + 1):
        row = [ws.cell(row=i, column=j).value for j in range(1, ws.max_column + 1)]
        vals = [str(c).strip().lower() if c else "" for c in row]
        if vals[0] in ("#", "fase") or any(h in vals for h in ["nome", "região", "regiao", "descrição", "gps", "espécie"]):
            headers = [str(c).strip() if c else f"col_{j}" for j, c in enumerate(row)]
            return i, headers
    return 3, []


def find_col(headers, *patterns):
    for i, h in enumerate(headers):
        h_lower = h.lower().strip()
        for p in patterns:
            if p.lower() in h_lower:
                return i
    return -1


def process_sheet(ws, sheet_name, category_id):
    header_row, headers = find_header_row(ws)
    if not headers:
        row3 = [ws.cell(row=3, column=j).value for j in range(1, ws.max_column + 1)]
        headers = [str(c).strip() if c else f"col_{i}" for i, c in enumerate(row3)]
        header_row = 3

    if not headers:
        return []

    name_col = find_col(headers, "nome", "poi", "prato", "doce", "espécie", "especie",
                         "festa", "spot", "produtor", "castelo", "obra", "miradouro",
                         "linha", "sítio", "sitio", "quinta", "etapa", "rota")
    if name_col < 0:
        name_col = 1

    region_col = find_col(headers, "região", "regiao", "sub-região")
    desc_col = find_col(headers, "descrição", "descricao")
    address_col = find_col(headers, "morada", "cidade", "localidade", "localização")
    website_col = find_col(headers, "website", "web", "referência", "visita")
    contact_col = find_col(headers, "contacto", "contato", "email")
    price_col = find_col(headers, "preço", "preco")
    hours_col = find_col(headers, "horário", "horario")
    rarity_col = find_col(headers, "raridade", "nível pérola", "qualidade")

    # GPS columns - check headers for GPS-like columns
    gps_cols = []
    for i, h in enumerate(headers):
        h_lower = h.lower()
        if any(k in h_lower for k in ["gps", "trilho no maps", "rota no maps", "mapa"]):
            gps_cols.append(i)

    # Also check latitude/longitude separate columns
    lat_col = find_col(headers, "latitude")
    lng_col = find_col(headers, "longitude")

    items = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_values = [ws.cell(row=row_idx, column=j).value for j in range(1, ws.max_column + 1)]
        
        if not any(c for c in row_values):
            continue

        name_val = row_values[name_col] if name_col < len(row_values) else None
        # Handle HYPERLINK in name column
        name = str(name_val).strip() if name_val else None
        if name and name.startswith("=HYPERLINK"):
            # Extract display text from HYPERLINK formula
            m = re.search(r'"([^"]*)"[,)]\s*"([^"]*)"', name)
            name = m.group(2) if m else None

        if not name or name == "None" or len(name) < 2:
            continue
        if re.match(r'^\d+\.\s+[A-Z]', name) and "(" in name and ")" in name:
            continue

        # Region
        region_val = str(row_values[region_col]).strip() if region_col >= 0 and region_col < len(row_values) and row_values[region_col] else ""
        if region_val.startswith("=HYPERLINK"):
            region_val = ""
        region = normalize_region(region_val)

        # Description
        desc_val = row_values[desc_col] if desc_col >= 0 and desc_col < len(row_values) else None
        description = str(desc_val).strip() if desc_val and str(desc_val) != "None" else ""
        if description.startswith("=HYPERLINK"):
            description = ""

        # GPS - try multiple approaches
        location = None
        
        # 1. Separate lat/lng columns
        if lat_col >= 0 and lng_col >= 0:
            try:
                lat = float(row_values[lat_col]) if row_values[lat_col] else None
                lng = float(row_values[lng_col]) if row_values[lng_col] else None
                if lat and lng and -90 <= lat <= 90 and -180 <= lng <= 180:
                    location = {"lat": lat, "lng": lng}
            except (ValueError, TypeError):
                pass

        # 2. HYPERLINK formulas in GPS columns
        if not location:
            for gc in gps_cols:
                cell_val = row_values[gc] if gc < len(row_values) else None
                if cell_val:
                    cell_str = str(cell_val)
                    if "HYPERLINK" in cell_str or "maps" in cell_str.lower():
                        location = extract_gps_from_hyperlink(cell_str)
                        if location:
                            break
                    else:
                        location = extract_gps_plain(cell_str)
                        if location:
                            break

        # 3. Scan ALL cells for HYPERLINK with maps coordinates
        if not location:
            for ci in range(len(row_values)):
                cv = row_values[ci]
                if cv and "HYPERLINK" in str(cv) and "maps" in str(cv).lower():
                    location = extract_gps_from_hyperlink(str(cv))
                    if location:
                        break

        # Address
        address = ""
        if address_col >= 0 and address_col < len(row_values) and row_values[address_col]:
            a = str(row_values[address_col]).strip()
            if a != "None" and not a.startswith("=HYPERLINK"):
                address = a

        # Metadata
        metadata = {}
        for label, col_idx in [("website", website_col), ("contact", contact_col),
                                ("price", price_col), ("hours", hours_col), ("rarity", rarity_col)]:
            if col_idx >= 0 and col_idx < len(row_values) and row_values[col_idx]:
                val = str(row_values[col_idx]).strip()
                if val != "None" and not val.startswith("=HYPERLINK"):
                    metadata[label] = val
                elif val.startswith("=HYPERLINK"):
                    m = re.search(r'HYPERLINK\("([^"]*)"', val)
                    if m:
                        metadata[label] = m.group(1)

        item = {
            "id": str(uuid.uuid4()),
            "name": name[:300],
            "description": description[:2000],
            "category": category_id,
            "subcategory": None,
            "region": region,
            "location": location,
            "address": address[:500],
            "image_url": None,
            "tags": [category_id],
            "metadata": metadata if metadata else None,
            "source_sheet": sheet_name,
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        if location:
            item["geo_location"] = {
                "type": "Point",
                "coordinates": [location["lng"], location["lat"]]
            }
        items.append(item)

    return items


def main():
    print("=" * 60)
    print("PORTUGAL VIVO — Universal Importer v2 (HYPERLINK GPS)")
    print("=" * 60)

    client = MongoClient(MONGO_URL)
    db = client[DB_NAME]

    db.heritage_items.delete_many({})
    db.routes.delete_many({})
    print("\nCleared existing data.\n")

    wb = openpyxl.load_workbook("PortugalVivo_v19.xlsx", read_only=False, data_only=False)

    total_imported = 0
    total_with_gps = 0
    category_counts = {}
    region_counts = {}

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        category_id = None
        for key, val in SHEET_CATEGORY_MAP.items():
            if key in sheet_name or sheet_name in key:
                category_id = val
                break
        if not category_id:
            print(f"⚠️  No mapping: {sheet_name}")
            continue

        ws = wb[sheet_name]
        items = process_sheet(ws, sheet_name, category_id)

        if items:
            db.heritage_items.insert_many(items)
            gps_count = sum(1 for i in items if i.get("location"))
            total_imported += len(items)
            total_with_gps += gps_count
            category_counts[category_id] = category_counts.get(category_id, 0) + len(items)
            for item in items:
                region_counts[item["region"]] = region_counts.get(item["region"], 0) + 1
            gps_pct = round(gps_count / len(items) * 100) if items else 0
            print(f"✅  {sheet_name} → {category_id}: {len(items)} items ({gps_count} GPS = {gps_pct}%)")
        else:
            print(f"❌  {sheet_name}: 0 items")

    wb.close()

    # Indexes
    db.heritage_items.create_index("category")
    db.heritage_items.create_index("region")
    db.heritage_items.create_index("name")
    db.heritage_items.create_index([("geo_location", "2dsphere")])

    print(f"\n{'='*60}")
    print(f"TOTAL: {total_imported} POIs ({total_with_gps} com GPS = {round(total_with_gps/total_imported*100)}%)")
    print(f"\n📂 Categorias ({len(category_counts)}):")
    for cat, count in sorted(category_counts.items(), key=lambda x: -x[1]):
        print(f"  {cat}: {count}")
    print(f"\n🗺️  Regiões ({len(region_counts)}):")
    for reg, count in sorted(region_counts.items(), key=lambda x: -x[1]):
        print(f"  {reg}: {count}")
    print("=" * 60)

    client.close()


if __name__ == "__main__":
    main()
